#!/usr/bin/env python3
"""Converte CSV/TSV/Excel (aba Banco de Dados) em JSON para importar no app."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path


ALIASES = {
    "name": {"name", "alimento", "alimentos", "food"},
    "portion": {"portion", "porcao", "porção", "porcao (g, ml ou qtd)", "porção (g, ml ou qtd)"},
    "carbs": {"carbs", "carboidratos", "carboidratos (g)", "carbohydrates"},
    "protein": {"protein", "proteinas", "proteínas", "proteínas (g)", "proteinas (g)"},
    "fat": {"fat", "gorduras", "gorduras (g)"},
    "satFat": {
        "satfat",
        "gorduras sat",
        "gorduras sat(g)",
        "gorduras saturadas",
        "gorduras saturadas (g)",
        "saturated fat",
    },
    "fiber": {"fiber", "fibras", "fibras (g)", "fibre"},
    "sodium": {"sodium", "sodio", "sódio", "sodio (mg)", "sódio (mg)"},
    "addedSugar": {"addedsugar", "acucares", "açucares", "açúcares", "acucar adicionado", "açúcar adicionado"},
    "netCarbs": {"netcarbs", "carboidratos liquidos", "carboidratos líquidos", "net carbs"},
}


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    s = s.replace("ú", "u").replace("ü", "u")
    s = s.replace("ç", "c")
    return re.sub(r"\s+", " ", s)


def slug(name: str, idx: int) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", norm(name)).strip("-")
    return (base[:48] or f"food-{idx}")


def to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def map_headers(headers: list[str]) -> dict[str, int]:
    mapped = {}
    for i, h in enumerate(headers):
        nh = norm(h)
        for key, aliases in ALIASES.items():
            if nh in aliases or any(a in nh for a in aliases if len(a) > 4):
                mapped.setdefault(key, i)
    return mapped


def row_to_food(row: list, cols: dict[str, int], idx: int) -> dict | None:
    if "name" not in cols:
        return None
    name = str(row[cols["name"]] if cols["name"] < len(row) else "").strip()
    if not name:
        return None

    def get(key, default=0.0):
        if key not in cols:
            return default
        i = cols[key]
        return to_float(row[i] if i < len(row) else None, default)

    carbs = get("carbs")
    fiber = get("fiber")
    net = get("netCarbs", max(carbs - fiber, 0))
    return {
        "id": slug(name, idx),
        "name": name,
        "portion": get("portion", 100) or 100,
        "carbs": round(carbs, 3),
        "protein": round(get("protein"), 3),
        "fat": round(get("fat"), 3),
        "satFat": round(get("satFat"), 3),
        "fiber": round(fiber, 3),
        "sodium": round(get("sodium"), 3),
        "addedSugar": round(get("addedSugar"), 3),
        "netCarbs": round(net, 3),
        "source": "import",
    }


def from_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        reader = csv.reader(f, dialect)
        rows = list(reader)
    if not rows:
        return []
    cols = map_headers(rows[0])
    start = 1
    if "name" not in cols:
        # assume fixed order without header
        cols = {
            "name": 0,
            "portion": 1,
            "carbs": 2,
            "protein": 3,
            "fat": 4,
            "satFat": 5,
            "fiber": 6,
            "sodium": 7,
            "addedSugar": 8,
            "netCarbs": 9,
        }
        start = 0
    foods = []
    for i, row in enumerate(rows[start:], start=1):
        food = row_to_food(row, cols, i)
        if food:
            foods.append(food)
    return foods


def from_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb["Banco de Dados"] if "Banco de Dados" in wb.sheetnames else wb.active
    # layout da planilha: headers na linha 2, dados a partir da 3, colunas C-J
    foods = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name:
            continue
        row = [
            name,
            ws.cell(r, 4).value,
            ws.cell(r, 5).value,
            ws.cell(r, 6).value,
            ws.cell(r, 7).value,
            ws.cell(r, 8).value,
            ws.cell(r, 9).value,
            ws.cell(r, 10).value,
            0,
            None,
        ]
        cols = {
            "name": 0,
            "portion": 1,
            "carbs": 2,
            "protein": 3,
            "fat": 4,
            "satFat": 5,
            "fiber": 6,
            "sodium": 7,
            "addedSugar": 8,
            "netCarbs": 9,
        }
        food = row_to_food(row, cols, r)
        if food:
            foods.append(food)
    return foods


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa alimentos para JSON do Dieta Tracker")
    parser.add_argument("input", type=Path, help="Arquivo .csv, .tsv ou .xlsx")
    parser.add_argument("-o", "--output", type=Path, default=Path("data/foods.imported.json"))
    args = parser.parse_args()

    suffix = args.input.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        foods = from_csv(args.input)
    elif suffix in {".xlsx", ".xlsm"}:
        foods = from_xlsx(args.input)
    else:
        raise SystemExit("Formato não suportado. Use csv/tsv/xlsx.")

    # dedupe by name
    seen = set()
    unique = []
    for f in foods:
        key = norm(f["name"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(f)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(unique, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Exportados {len(unique)} alimentos → {args.output}")


if __name__ == "__main__":
    main()
